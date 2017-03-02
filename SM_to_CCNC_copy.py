# coding: utf-8
import pandas as pd
import os
import numpy as np
import openpyxl
import argparse
import textwrap
# Set pandas options
pd.set_option('max_rows',5000)
pd.set_option('max_columns',5000)

def surveyMonkeyToCCNC(args):

    template_df = load_template(args.surveyTemplate) ## surveyTemplate (with question names)를 template_df로 지정
    real_df = excel_merge_rearrange(args.inputExcel) ## input으로 들어간 df들을 merge해서 real_df으로 지정
                                                     ## excel_merge_rearrange function 사용
    
    ## 아래의 load_template function과 excel_merge_rearrange function을 통해 각각 얻은 template_df와 real_df에 적용
    ## surveyTemplate의 question name을 real_df의 question_name으로 지정; real_df의 마지막 column에 question_name column data 입력
    real_df['question name'] = template_df['question name']
    
    ## template (with all data formula)를 python 환경 내에서 excel로 오픈 - openpyxl
    template_openpyxl = openpyxl.load_workbook(args.template, data_only=False)
                                                                              
    update_template_excel(template_openpyxl, real_df) ## update_template_excel function 참고; real_df의 값들을 template excel에 입력
    template_openpyxl.save(args.output) ## 결과를 "output"으로 save

    print '=========='
    print 'Completed'
    print '=========='


def load_template(survey_monkey_template):
    '''
    Takes input of survey monkey template location.
    A row of question name should be in the excel file.
    '''
    # Load survey monkey template with question names
    template_df = pd.read_excel(survey_monkey_template).T.reset_index() ## df.T - transpose index (=row) and columns
                                                                        ## df.reset_index() - return new df with labeling info
    template_df.columns = ['title','subtitle','question name','answers'] ## columns 지정
    template_df.ix[:10,'question name'] = 'personal_info' ## :10 (index)까지의 'question name'(column)에 'personal_info'를 labeling 
    return template_df

def excel_merge_rearrange(excelList):
    '''
    Takes input of list of excel file locations.

    It extracts the last subject's data in the excel List,
    concatenating them to a pandas dataframe.
    Also, the 'title' column is decoded to unicode.
    '''
    ## pd.concat([pd.read_excel(x) for x in excelList]) - excelList에 있는 excel들을 모두 불러와서 취합
    ## index_col=range(9) - able to pass the column number or column name you wish to use as the index -> 0~9 rows 지정
    ## axis=1 - along the columns; cf. axis=0 as default and along the rows
    all_sheets = pd.concat([pd.read_excel(x, index_col=range(11)) for x in excelList],
                           axis=1).reset_index()        
    
    # Extracts the last subject before transposing
    ## Survey_monkey를 이용해 데이터를 입력하면, 하나의 dataframe에 지속적으로 데이터 누적 -> new subject's data를 new row에 입력.
    ## all_sheets df의 index 중 0(question names)과 -1(last subject's data)에 해당하는 rows만 extract 함.
    ## transpose & reset index
    real_df = all_sheets.ix[all_sheets.index[[0,-1]],:].T.reset_index()
    
    #real_df = all_sheets.T.reset_index()
    ## columns 이름 지정
    real_df.columns = ['title','subtitle','answers']
    
    ## title 중에 css로 표기된 부분을 제거
    real_df.title = real_df.title.str.replace('<[^>]+>','')
    ## title의 속성을 unicode로 변환(?)
    real_df.title = real_df.title.map(unicode)
    return real_df


def update_template_excel(template_openpyxl, real_df):
    ## enumerate - index 값을 주어, list에 순서를 부여
    ## normally used in for loop as below
    ## for문처럼 반복되는 구간에서 객체가 현재 어느 위치에 있는지 알려주는 인덱스 값이 필요할 때 사용
    ## 아래에서는 num이 index number를, question_name이 list 내의 items를 지정
    for num, question_name in enumerate(template_openpyxl.sheetnames[:-5]): 
        # 뒤에서 5개(결과지)를 제외한 템플레이트의 모든 sheet를 looping. 
        # 엑셀파일에 데이터가 바로 들어가는 것이 아니고, 행 열, 4칸 3칸씩 띄우고 데이터가 들어감
        startRow = 5
        startCol = 4
        
        
        ## template excel의 각 sheet의 이름(question_name)을 차용해 worksheet를 불러옴.
        sheet = template_openpyxl.get_sheet_by_name(question_name) 
        
        # 서베이멍키에서 불러온 자료를 불러옴
        ## 각 question_name에 대해 loop이 돌아가며, 각각에 맞는 데이터들을 survey_monkey real_df에서 찾음
        ## real_df['question name'].str.strip()==question_name -> 각 question_name에 맞는 것들을 찾아줌
        ## real_df.ix[Above] -> 각 question_name에 맞는 것들에 대해, 각각 indexing을 해주고, value(eg. answers)를 찾음
        data = real_df.ix[real_df['question name'].str.strip()==question_name]
        
        if question_name == u'기본정보':
            startRow -= 1
            startCol -= 1

            info_list = [u'병록번호 (입력 불필요)', u'이름', 
                         u'응답자 ID', u'시작 날짜', 
                         u'IP 주소', u'컬렉터 ID', u'종료 날짜', u'E/O (입력 불필요)']

            for info, answer in [(x,str(real_df.ix[real_df.title==x,'answers'].get_values()[0])) for x in info_list]:
                if u'날짜' in info:
                    answer = answer[:10]
                elif u'E/O (입력 불필요)' in info:
                    answer = int(answer)
                try:
                    sheet.cell(row = startRow, column = startCol).value = answer
                except:
                    sheet.cell(row = startRow, column = startCol).value = ''
                startRow+=1

        elif question_name == '(H, I) SCL, EF': #이런 이름을 가진 sheet는
            for num,answer in enumerate(data.answers): #나머지를 프린트
                doubleCellWrite(sheet.cell, startRow, startCol, answer)
                if num==28:
                    startRow+=5 #다섯 칸 띄우고
                else:
                    startRow+=1

        ## ELSQ (FEP, OCD)
        elif question_name == '(8) ELSQ': #이런 이름을 가진 sheet는
            # 3, 4, 3, 4가 반복되는 리스트를 만듬 (column이 3,4 순으로 들어감)
            columns = [startCol, startCol+1] * (len(data.answers)/2)  ## len ???
            # make double numbered row list
            rows = sorted(range(startRow, len(data.answers)) * 2)

            for num, (row, column, answer) in enumerate(zip(rows, columns, data.answers)):
                sheet.cell(row=row, column=column).value = answer#데이터 입력

        ## ELSQ (DICC)
        elif question_name == '(22) ELSQ': #이런 이름을 가진 sheet는
            # 3, 4, 3, 4가 반복되는 리스트를 만듬 (column이 3,4 순으로 들어감)
            columns = [startCol, startCol+1] * (len(data.answers)/2)  ## len ???
            # make double numbered row list
            rows = sorted(range(startRow, len(data.answers)) * 2)

            for num, (row, column, answer) in enumerate(zip(rows, columns, data.answers)):
                sheet.cell(row=row, column=column).value = answer#데이터 입력                
                
        ## K-SFS 추가 (FEP, OCD용)
        elif question_name == '(K) K-SFS':
            data_SFS = data[1:72]
            for answer in data_SFS.answers:
                doubleCellWrite(sheet.cell, startRow, startCol, answer)
                startRow +=1
                
        ## K-SFS 추가 (DICC용)
        elif question_name == '(12) K-SFS':
            data_SFS = data[1:72]
            for answer in data_SFS.answers:
                doubleCellWrite(sheet.cell, startRow, startCol, answer)
                startRow +=1
        
        ## K-SFS-보호자
        elif question_name == '(K) K-SFS-P':
            data_SFS_P = data[3:74]
            for answer in data_SFS_P.answers:
                doubleCellWrite(sheet.cell, startRow, startCol, answer)
                startRow +=1
        
        ## YBOCS 추가 (OCD)        
        elif question_name == '(12) YBOCS (Self)':
            data_YBOCS = data[-12:]
            for answer in data_YBOCS.answers:
                doubleCellWrite(sheet.cell, startRow, startCol, answer)
                startRow +=1
        
        ## SFRT (DICC)
        elif question_name == '(1)SFRT':
            data_SFRT = data[29:]
            startCol -= 1
            startRow -= 2
            for num,answer in enumerate(data_SFRT.answers):            
                sheet.cell(row = startRow, column = startCol).value = answer
                
                if num == 13 or num == 27 or num == 42 or num == 56 or num == 71 or num == 85 or num == 100 or num == 114 or num == 129 or num == 143 or num == 158 or num == 172 or num == 187 or num == 201 or num == 216 or num == 230 or num == 245 or num == 259:
                    startRow -=13
                    startCol +=1
                elif num == 28 or num == 57 or num == 86 or num == 115 or num == 144 or num == 173 or num == 202 or num == 231 or num == 260:
                    startRow +=14
                    startCol -=2
                else:
                    startRow +=1

        ## IPSAQ (DICC & NOR)            
        elif question_name == '(2)IPSAQ' or question_name == '(15)IPSAQ':
            startRow -=3
            startCol -=2
            for num,answer in enumerate(data.answers):
                if not (num % 2 == 0):
                    doubleCellWrite_IPSAQ(sheet.cell, startRow, startCol, answer)
                    startRow +=1
        
        ## PQ-B (DICC & NOR)
        elif question_name == '(3)PQ-B' or question_name == '(16)PQ-B':
            for num,answer in enumerate(data.answers):
                if (num % 2 == 0):
                    doubleCellWrite(sheet.cell, startRow, startCol, answer)
                    startRow +=1            
                    
        ## BPSS-AS-P (DICC)
        elif question_name == '(25) BPSS-AS-P':
            for num,answer in enumerate(data.answers):
                doubleCellWrite(sheet.cell, startRow, startCol, answer)
                if num == 13:
                    startRow +=3
                else:
                    startRow +=1
                    
        else:
            for answer in data.answers:# 서베이멍키에서 불러온 각 문제의 자료들을 looping
                doubleCellWrite(sheet.cell, startRow, startCol, answer)
                startRow +=1

                
def doubleCellWrite(cell, row, column, answer):
    cell(row = row, column = column).value = answer
    cell(row = row, column = column+1).value = answer

def doubleCellWrite_IPSAQ(cell, row, column, answer):
    cell(row = row, column = column).value = answer
    cell(row = row, column = column+11).value = answer

if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description=textwrap.dedent('''\
            {codeName} : Saves the Survey monkey data to CCNC format
            ========================================
            '''.format(codeName=os.path.basename(__file__))))

    parser.add_argument(
        '-i', '--inputExcel',
        help='Survey Monkey exported data files',
        nargs='+',
        default=os.getcwd())

    parser.add_argument(
        '-g', '--group',
        help='Group fo the subject',
        )

    parser.add_argument(
        '-t', '--template',
        help='Excel template with all formulas',
        default = '/Users/takwan_kim/survey_monkey/SM_code/FEP_template.xlsx')

    parser.add_argument(
        '-st', '--surveyTemplate',
        help='Survey monkey template with question name column',
        default = '/Users/takwan_kim/survey_monkey/SM_code/FEP_merged_revised.xlsx')

    parser.add_argument(
        '-o', '--output',
        help='Excel ouput',
        default = 'out.xlsx')

    args = parser.parse_args()


    surveyMonkeyToCCNC(args)
