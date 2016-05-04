# coding: utf-8
import pandas as pd
import os
import numpy as np
import openpyxl
import argparse
import textwrap


# Set pandas options
pd.set_option('max_rows',5000)
pd.set_option('max_columns',5000)

def main(args):
    # import merged template
    group_template_df = pd.read_excel(args.groupTemplate)
    template_df = group_template_df.T.reset_index()
    template_df.columns = ['title','subtitle','question name','answers']
    template_df.ix[:8,'question name'] = 'personal_info'


    # Survey monkey data
    all_sheets = pd.concat([pd.read_excel(x) for x in args.inputExcel],axis=1)
    df = all_sheets.T.reset_index()
    df.columns = ['title','subtitle','answers']
    df.title = df.title.str.replace('<[^>]+>','')
    df.title = df.title.map(unicode)


    # template df 에서 question name 가지고오기
    df['question name'] = template_df['question name']


    # Formula template
    template_openpyxl = openpyxl.load_workbook(args.template, data_only=False)
    #out_excel = openpyxl.load_workbook('/Users/kangik/Desktop/xlrd_prac.xls', data_only=False)

    #print df.head(10)
    for num, sheet_name in enumerate(template_openpyxl.sheetnames[:-5]): # 뒤에서 5개를 제외한 템플레이트의 모든 sheet를 looping
        # 엑셀파일에 데이터가 바로 들어가는 것이 아니고,
        # 행 열, 4칸 3칸씩 띄우고 데이터가 들어감
        startRow = 5
        startCol = 4

        sheet = template_openpyxl.get_sheet_by_name(sheet_name) # sheet를 불러옴
        data = df.ix[df['question name'].str.strip()==sheet_name] # 서베이멍키에서 불러온 자료를 불러옴
        
        if sheet_name == u'기본정보':
            startRow -= 1
            startCol -= 1
            for answer in [df.ix[df.title == u'응답자 ID', 'answers'].get_values()[0],
                            str(df.ix[df.title == u'성', 'answers'].get_values()[0])+ str(df.ix[df.title == u'이름', 'answers'].get_values()[0]),
                            df.ix[df.title == u'사용자 정의 데이터', 'answers'].get_values()[0],
                            str(df.ix[df.title == u'시작 날짜', 'answers'].get_values()[0])[:10],
                            df.ix[df.title == u'IP 주소', 'answers'].get_values()[0],
                            df.ix[df.title == u'컬렉터 ID', 'answers'].get_values()[0],
                            str(df.ix[df.title == u'종료 날짜', 'answers'].get_values()[0])[:10]]:
                print answer
                sheet.cell(row = startRow, column = startCol).value = str(answer)
                startRow+=1


        elif sheet_name == '(H, I) SCL, EF': #이런 이름을 가진 sheet는
            for num,answer in enumerate(data.answers[:29]): #처음 29줄을 프린트하고
                sheet.cell(row = startRow, column=startCol).value =  answer
                sheet.cell(row=startRow, column=startCol+1).value = answer
                startRow+=1
            startRow+=4 #세칸 띄우고
            for num,answer in enumerate(data.answers[29:]): #나머지를 프린트
                sheet.cell(row=startRow, column=startCol).value =  answer
                sheet.cell(row=startRow, column=startCol+1).value = answer
                startRow+=1

        elif sheet_name == '(8) ELSQ': #이런 이름을 가진 sheet는
            columns = [startCol, startCol+1]*(len(data.answers)/2) #3, 4, 3, 4가 반복되는 리스트를 만듬 (column이 3,4 순으로 들어감)
            initialC = 3
            for num, column in enumerate(columns):
                if column == 3 and data.answers.tolist()[num]==0: #column 3에 들어갈 자료가 0이면
                    sheet.cell(row=startRow, column=column).value = data.answers.tolist()[num] # 0을 쓰고
                    sheet.cell(row=startRow, column=column+1).value = '' # 그 옆에는 '' 빈칸을 넣는다
                elif column == 3 and answer!=0: #column 3에 들어갈 자료가 0이 아니면
                    sheet.cell(row=startRow, column=column).value = data.answers.tolist()[num] 
                    sheet.cell(row=startRow, column=column+1).value = data.answers.tolist()[num+1] # 그 옆에 자료를 쓴다

                if column > initialC:
                    startRow+=1 # 2 data point가 지날 때마다 한줄씩 증가됨

        else:
            for answer in data.answers:# 서베이멍키에서 불러온 각 문제의 자료들을 looping
                sheet.cell(row=startRow, column=startCol).value = answer#데이터 입력
                sheet.cell(row=startRow, column=startCol+1).value = answer#더블체크 칸 데이터 동일하게 입력
                startRow +=1

    template_openpyxl.save(args.output)
    print 'Completed'



if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description=textwrap.dedent('''\
            {codeName} : Saves the Survey monkey data to CCNC format
            ========================================
            '''.format(codeName=os.path.basename(__file__))))

    parser.add_argument(
        '-i', '--inputExcel',
        help='Survey Monkey export file',
        nargs='+',
        default=os.getcwd())

    parser.add_argument(
        '-g', '--group',
        help='Group fo the subject',
        )

    parser.add_argument(
        '-t', '--template',
        help='Excel template with all formulas',
        default = '/ccnc_bin/survery_monkey/template.xlsx')

    parser.add_argument(
        '-gt', '--groupTemplate',
        help='Survey monkey template with question name column',
        default = '/ccnc_bin/survery_monkey/merged.xlsx')

    parser.add_argument(
        '-o', '--output',
        help='Excel ouput',
        default = 'out.xlsx')

    args = parser.parse_args()


    main(args)
